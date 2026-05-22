import os
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

FILE_INPUT = "Hasil_Ekstrak_Rincian_ZN_temp.xlsx"
FILE_OUTPUT = "Hasil_Ekstrak_Rincian_ZN.xlsx"

BULAN_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MEI": 5, "JUN": 6,
    "JUL": 7, "AGU": 8, "SEP": 9, "OKT": 10, "NOV": 11, "DES": 12,
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "Mei": 5, "Jun": 6,
    "Jul": 7, "Agu": 8, "Sep": 9, "Okt": 10, "Nov": 11, "Des": 12
}

def parse_ke_datetime(val_tgl):
    if val_tgl is None:
        return None
    if isinstance(val_tgl, datetime):
        return val_tgl
    
    tgl_str = str(val_tgl).strip()
    try:
        if "-" in tgl_str:
            return datetime.strptime(tgl_str.split()[0], "%Y-%m-%d")
            
        parts = tgl_str.split()
        if len(parts) == 3:
            day = int(parts[0])
            month = BULAN_MAP[parts[1]]
            year = int(parts[2])
            return datetime(year, month, day)
    except:
        pass
    return None

if not os.path.exists(FILE_INPUT):
    print(f"--> File target {FILE_INPUT} tidak ditemukan!")
    exit()

print(f"--> Membuka file: {FILE_INPUT}")
wb_asal = openpyxl.load_workbook(FILE_INPUT, data_only=True)
sheet_asal = wb_asal.active

wb_baru = openpyxl.Workbook()
sheet_baru = wb_baru.active
sheet_baru.title = sheet_asal.title

header_row = [cell.value for cell in sheet_asal[1]]
header_row.append("JT")
sheet_baru.append(header_row)

idx_tgl_input = 0
idx_nama = 1
idx_toko = 2
idx_jatuh_tempo = 11

last_group_key = None
hari_ini = datetime.now()

print("--> Menyusun baris data dan menghitung umur Jatuh Tempo (JT)...")

for row in sheet_asal.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
        
    current_group_key = f"{str(row[idx_tgl_input]).strip()}||{str(row[idx_nama]).strip()}||{str(row[idx_toko]).strip()}"
    
    if last_group_key is not None and current_group_key != last_group_key:
        sheet_baru.append([])
        
    jt_val_raw = row[idx_jatuh_tempo]
    dt_jt = parse_ke_datetime(jt_val_raw)
    
    if dt_jt:
        selisih_hari = (hari_ini - dt_jt).days
        teks_jt = f"{selisih_hari} Hari"
    else:
        teks_jt = ""
        
    new_row_data = list(row) + [teks_jt]
    sheet_baru.append(new_row_data)
    
    last_group_key = current_group_key

print("--> Menghitung lebar kolom otomatis (Auto-fit)...")
for col in sheet_baru.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))
    sheet_baru.column_dimensions[col_letter].width = max(max_len + 3, 10)

wb_baru.save(FILE_OUTPUT)
wb_asal.close()
wb_baru.close()

print(f"--> Selesai! File '{FILE_OUTPUT}' telah di tambah pembatas baris kosong dan kolom JT.")